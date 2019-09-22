# Python script to create M3U8 file from excell (xlsx)

"""
#mport os.pathi
#os.path.join('/my/root/directory', 'in', 'here')
'/my/root/directory/in/here'
"""




# function to detect the current working folder and create output:
def working_folder_path_detection():

    # Import the external libraries:
    import os
    import shutil

    # Global variables:
    global current_working_directory_path
    global input_directory_path
    global output_directory_path

    print("\nWorking folder path detection and structure creation function:\n")

    # Check current working directory:

    current_python_working_directory = os.getcwd()
    print("Current python working directory: " + str(current_python_working_directory))

    # Go up in the folder structure:
    os.chdir("../")

    # Detect the current working directory:
    current_working_directory_path = os.getcwd()
    print("Current working directory: " + str(current_working_directory_path))



    # Create the input folder if not exists:
    if not os.path.exists(str(input_directory_name)):
        os.makedirs(input_directory_name)

    # Check the input folder path:
    os.chdir(str(input_directory_name))
    input_directory_path = os.getcwd()
    print("Current input directory path: " + str(input_directory_path))


    # Go up in the folder structure back to the working directory:
    os.chdir("../")



    # Delete the output folder for the M3U and JSON files if exists then create a new one:
    if os.path.exists(str(output_directory_name)):
        shutil.rmtree(str(output_directory_name), ignore_errors=False, onerror=None)
        os.makedirs(output_directory_name)

    # Create a new folder in case of the output is missing:
    else:
        os.makedirs(output_directory_name)

    # Go to the output folder:
    os.chdir(str(output_directory_name))

    # Check the output directory path:
    output_directory_path = os.getcwd()
    print("Current output directory path: " + str(output_directory_path))

    # End of working folder detection and structure creation function:

# Read the excel file and create M3U files:
def excell_to_M3U_files():

    # Import external libraries
    import openpyxl
    import os.path
    import codecs
    import os

    # Define the M3U file extension:
    m3u_file_extension = '.m3u'

    # Define the M3U file header information:
    m3u_extinf_content = 'EXTINF:0,'

    print('Current excel file name: ' + input_excel_file_name)

    input_excell_file_path = os.path.join(input_directory_path, input_excel_file_name)
    print('Input excel file path: ' + input_excell_file_path)

    # Change working directory to the input folder in order to open the excel file
    os.chdir(str(input_directory_path))

    # To open the workbook workbook object is created
    workbook = openpyxl.load_workbook(input_excel_file_name)

    # Get workbook active current_sheet object
    current_sheet = workbook.active

    # Change the working directory to Output in order to write the M3U files:
    os.chdir(str(output_directory_path))

    # Set the loop initiator to 2 because the first row of the excel file does not contain any relevant data related to channel information
    current_raw = 2

    while True:

        # Print the start of the function
        print('M3U file creation has been started.\n')

        webradio_name = current_sheet.cell(row=current_raw, column=1)
        webradio_stream = current_sheet.cell(row=current_raw, column=2)

        if (webradio_name != None) and (webradio_stream.value != None) :

            current_m3u_extinf = str(m3u_extinf_content) + str(webradio_name.value) + str('\n')
            current_m3u_stream = str(webradio_stream.value)
            current_m3u_filename = str(webradio_name.value) + str(m3u_file_extension)

            # print(current_m3u_extinf)
            # print(current_m3u_stream)
            # print(current_m3u_filename)
            # Open the file

            m3u_file = codecs.open(str(current_m3u_filename), "w", "utf-8")
            m3u_file.write(u'\ufeff')
            m3u_file.write("#EXTM3U\n")
            m3u_file.write(current_m3u_extinf)
            m3u_file.write(current_m3u_stream)
            m3u_file.write("\n")

            # Close the file
            m3u_file.close()

            # Increment the loop iterator then break at the end

        else:
            print("Blank cell detected at line: " + str(current_raw) + '\n')

        current_raw += 1
        if current_raw >= last_raw:
            break

    # End of read the excel file and create M3U files

# Read the excel file and create common M3U file:
def excell_to_common_M3U_file():

    # Import external libraries
    import openpyxl
    import os.path
    import codecs
    import os

    # Define the M3U file extension:
    m3u_file_extension = '.m3u'

    # Define the M3U file header information:
    m3u_extinf_content = 'EXTINF:0,'

    print('Current excel file name: ' + input_excel_file_name)

    input_excell_file_path = os.path.join(input_directory_path, input_excel_file_name)
    print('Input excel file path: ' + input_excell_file_path)

    # Change working directory to the input folder in order to open the excel file
    os.chdir(str(input_directory_path))

    # To open the workbook workbook object is created
    workbook = openpyxl.load_workbook(input_excel_file_name)

    # Get workbook active current_sheet object
    current_sheet = workbook.active

    # Change the working directory to Output in order to write the M3U files:
    os.chdir(str(output_directory_path))

    # Print the start of the function
    print('Common M3U file creation has been started.\n')

    # Set the loop initiator to 2 because the first row of the excel file does not contain any relevant data related to channel information
    current_raw = 2

    m3u_file = codecs.open(str(output_M3U_file_name), "w", "utf-8")
    m3u_file.write(u'\ufeff')
    m3u_file.write("#EXTM3U\n")
    m3u_file.write("\n")

    while True:

        webradio_name = current_sheet.cell(row=current_raw, column=1)
        webradio_stream = current_sheet.cell(row=current_raw, column=2)

        if (webradio_name != None) and (webradio_stream.value != None) :

            current_m3u_extinf = str(m3u_extinf_content) + str(webradio_name.value) + str('\n')
            current_m3u_stream = str(webradio_stream.value) + str('\n')

            # Write a new extinf line to the actual channel
            m3u_file.write(current_m3u_extinf)

            # Write a new stream line to the actual channel
            m3u_file.write(current_m3u_stream)

            # Write a new empty line to the actual channel
            m3u_file.write("\n")



        else:
            print("Blank cell detected at line: " + str(current_raw) + '\n')

        # Increment the loop iterator then break at the end
        current_raw += 1
        if current_raw >= last_raw:
            break

    # Close the file
    m3u_file.close()
    # End of read the excel file and create common M3U file

# Read the excel file and create JSON file to Ka-Radio:
def excel_to_JSON_file():

    # import the used external libraries
    import json
    from urllib.parse import urlparse
    import openpyxl
    import codecs
    import os.path



    print('Current excel file name: ' + input_excel_file_name)
    
    input_excell_file_path = os.path.join(input_directory_path, input_excel_file_name)
    print('Input excel file path: ' + input_excell_file_path)

    # Change working directory to the input folder in order to open the excel file
    os.chdir(str(input_directory_path))

    # To open the workbook workbook object is created
    workbook = openpyxl.load_workbook(input_excel_file_name)

    # Get workbook active current_sheet object
    current_sheet = workbook.active

    # Change working directory to the output folder in order to write the output JSON file
    os.chdir(str(output_directory_path))

    # Open the json file as UTF-8
    json_file = codecs.open(output_json_file_name, "w", "utf-8")

    # Print the start of the function
    print('JSON file creation has been started.\n')
    json_file.write(u'\ufeff')

    # Set the loop iterator to 2 because the first row does not contain any relevant data
    current_raw = 2

    while True:

        # Read the excel file data to variables
        current_json_name = current_sheet.cell(row=current_raw, column=1)
        current_json_stream = current_sheet.cell(row=current_raw, column=2)
        current_json_volume = current_sheet.cell(row=current_raw, column=3)

        # Check if any empty cell is present in the actual raw, if not continue
        if ( current_json_name.value != None ) and (current_json_stream.value != None) and (current_json_volume.value != None):

            # Print the actual row:
            print('actual excel line: ' + str(current_raw) + '\n')

            # Print the actual radio station name:
            print('actual radio station name: ' + str(current_json_name.value) + '\n')

            # Print the actual url:
            print('actual url: ' + str(current_json_stream.value) + '\n')

            # Parse the actual url then print its parts:


            # Print the actual parsed elements of the stream:
            print('actual parsed stream: \n')

            current_url = urlparse(str(current_json_stream.value))
            print('scheme: ' + str(current_url.scheme) + '\n')
            print('hostname: ' + str(current_url.hostname) + '\n')
            print('path: ' + str(current_url.path) + '\n')

            # If no port found then set to 80:
            parsed_port = str(current_url.port)
            substring = "None"
            if substring in parsed_port:
                parsed_port = '80'

            # Print the port:
            print("port: " + str(parsed_port) + "\n")

            # Give values to python string:
            current_python_string = {
                "Name": current_json_name.value,
                "URL": str(current_url.hostname),
                "File": str(current_url.path),
                "Port": parsed_port,
                "ovol": str(current_json_volume.value)
            }

            # Print the python string:
            print("python string: " + str(current_python_string) + "\n")

            # Convert python string into JSON string with separators , and :
            current_json_string = json.dumps(current_python_string, ensure_ascii=False, separators=(","":"))

            # Print the JSON string:
            print("JSON string: " + str(current_json_string) + "\n")

            # Write the JSON string into file:
            json_file.write(str(current_json_string) + '\n')

        # If any empty cell detected print an error message related to actual row
        else:
            print("Blank cell detected at line: " + str(current_raw) + '\n')

        # Increment the loop iterator then break at the end if the loop iterator higher than the last raw
        current_raw += 1
        if current_raw >= last_raw:
            break

    # Close the json file
    json_file.close()

    output_json_file_path = os.path.join(output_directory_path, output_json_file_name)
    print('Output JSON file path: ' + output_json_file_path)

    # Print the end of the function
    print('JSON file creation has been finished successfully.\n')

    # End of read the excel file and create JSON file to Ka-Radio


# Import external libraries
import time
import datetime
import os

# Install openpyxl package
os.system('pip install openpyxl')

# Define variables

# Give the location of the file
input_excel_file_name = 'Webradio.xlsx'

# Define the input folder name:
input_directory_name = "Input"

# Define the output folder name:
output_directory_name = "Output"

# Define the output JSON file name:
output_json_file_name = 'Webradio.txt'

# Define the output JSON file name:
output_M3U_file_name = 'Webradio.M3U'



# Print program information:
print("Python script to create M3U8 and JSON files from excell (xlsx)")
print("Written by dexter 2009.09.21")
current_date_and_time = datetime.datetime.now()
print('Current day and time: ' + str(current_date_and_time))

# define the last possible radio station:
last_raw = 256
print('Last raw in the excell file: ' + str(last_raw))

working_folder_path_detection()
excel_to_JSON_file()
excell_to_M3U_files()
excell_to_common_M3U_file()



# Sleep for 1 seconds
time.sleep(1)

